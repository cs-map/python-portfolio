# From an Excel workbook, record the associated Code Attachment title
# found in an ArcGIS Enterprise Portal.
# Use of Selenium requires administrative login for each scrape.

from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException
import openpyxl

# define your variables
wbFileName = ""  # path to workbook
sheet_name = ""  # worksheet name within workbook
column = ""  # column letter where urls are stored
admin_username = ""  # username for the arcgis user
admin_password = ""  # password for the arcgis user
path = "msedgedriver.exe"  # path where Edge driver is stored

# open the workbook where the app info is stored
wb = openpyxl.load_workbook(wbFileName, data_only=True)
ws = wb[sheet_name]
print("Worksheet loaded")

# identify the apps' urls
sheetUrls = ws[column]
codeUrls = []
for u in sheetUrls:
    text = u.value
    if text.startswith("https://"):
        codeUrls.append(text)
print("Urls ready")

# grab every app's Code Attachment title
for index, c in enumerate(codeUrls):
    # open the app Settings page in Microsoft Edge
    service = Service(executable_path=path)
    driver = webdriver.Edge(service=service)
    driver.get(c)
    print("Getting the Code Attachment name for app #{index}")
    try:
        # access Sign In frame
        locator = "oAuthFrame"
        element = ec.presence_of_element_located((By.ID, locator))
        WebDriverWait(driver, 10).until(element)
        frame = driver.find_element(By.ID, locator)
        driver.switch_to.frame(frame)
        # Sign In via ARCGIS (not single sign on)
        locator = "ago_Name"
        element = ec.presence_of_element_located((By.ID, locator))
        WebDriverWait(driver, 10).until(element)
        arcgis = driver.find_element(By.ID, locator)
        arcgis.click()
        # enter admin credentials
        username = driver.find_element(By.ID, "user_username")
        username.send_keys(admin_username)
        password = driver.find_element(By.ID, "user_password")
        password.send_keys(admin_password)
        signin = driver.find_element(By.ID, "signIn")
        signin.submit()
        # grab Code Attachment name
        locator = "span.label:nth-child(2)"
        element = ec.presence_of_element_located((By.CSS_SELECTOR, locator))
        WebDriverWait(driver, 10).until(element)
        codeTitle = driver.find_element(By.CSS_SELECTOR, locator)
        codeAttachment = codeTitle.text
    except TimeoutException:
        # mark a Code Attachment title that couldn't be grabbed
        ws.cell(row=index + 2, column=10).value = "skipped"
    else:
        # write a found Code Attachment title to the workbook
        ws.cell(row=index + 2, column=10).value = codeAttachment
    finally:
        # close the Microsoft Edge window
        driver.quit()
# save the workbook
wb.save(wbFileName)
