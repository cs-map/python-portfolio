import os
import re
import openpyxl
import numpy
from zipfile import ZipFile

# enter values for these items
downloadFolder = ""  # where the Code Attachment packages will be stored
newFolder = ""  # where the Code Attachments will be processed
wbFileName = ""  # Excel workbook containing Code Attachment urls
sheet_name = ""  # worksheet within the workbook where the urls are stored

# define variables and load the worksheet
downloadList = os.listdir(downloadFolder)
newList = os.listdir(newFolder)
wb = openpyxl.load_workbook(wbFileName, data_only=True)
ws = wb[sheet_name]
print("Workbook loaded")

# find the unique widgets in every Code Attachment
listOfWidgets = []
for f in newList:
    fullZipName = os.path.join(newFolder, f)
    fileName = r"config.json"
    fullFileName = os.path.join(fullZipName, fileName)
    with ZipFile(fullZipName) as my_zip:
        with my_zip.open(fileName) as my_file:
            jsonContents = my_file.read().decode('utf-8')
    searchPattern = re.compile(r"(widgets/)(\w+)(/)")
    allMatch = searchPattern.findall(jsonContents)
    listOfWidgets.append(f[:-4])
    matches = []
    for x in allMatch:
        matches.append(x[1])
    listOfWidgets.append(matches)
print("All widgets found")

# create a new sheet in the workbook to report the widgets
wsWidget = wb.create_sheet("Widgets", 0)
titles = listOfWidgets[::2]
widgets = listOfWidgets[1::2]
for index, t in enumerate(titles):
    wsWidget.cell(row=index+1, column=1).value = t
for index, w in enumerate(widgets):
    w = str(numpy.unique(w)).replace("'", "")
    wsWidget.cell(row=index+2, column=1).value = w
wb.save(wbFileName)
print("Analysis complete")
